from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from adminInventario2.models import *
import matplotlib
import json
from django.contrib.auth.decorators import login_required
from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum, F
from datetime import datetime, timedelta
from django.utils import timezone
from django.shortcuts import render
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

matplotlib.use("Agg")


@login_required
@csrf_exempt
def guardarVentas(request):
    if request.method == "POST":
        try:
            username = request.user.username
            # first_name = request.user.first_name
            # last_name = request.user.last_name
            # email = request.user.email

            # print(username, first_name, last_name, email)

            data = json.loads(request.body)
            fecha = data.get("fecha")
            cantidad = data.get("cantidad_productos")
            sub_precio = data.get("sub_precio")
            descuento = data.get("descuento")
            tipo_descuento = data.get("tipo_descuento")
            total_precio = data.get("total_precio")
            numero_orden = data.get("numero_orden")

            # Guardar los datos en la base de datos
            guardar_venta = Ventas(
                empleado=username,
                fecha_creacion=fecha,
                cantidad_productos=cantidad,
                sub_precio_venta=sub_precio,
                descuento=descuento,
                tipo_descuento=tipo_descuento,
                total_precio_venta=total_precio,
                numero_orden=numero_orden,  # Usar el número de orden proporcionado
            )

            guardar_venta.save()

            return JsonResponse(
                {
                    "mensaje": "Venta guardada correctamente",
                    "numero_orden": guardar_venta.numero_orden,
                }
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al procesar el JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
def guardarDetallesV(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            numero_orden = data.get("pedidoNumero")
            nombre_cliente = data.get("nombreCliente")
            correo_cliente = data.get("correoCliente")
            pedido = data.get("pedido")
            cantidad_productos = data.get("cantidad_productos")

            # Obtener la venta correspondiente al número de orden
            venta = get_object_or_404(Ventas, numero_orden=int(numero_orden))

            # Crear un detalle de venta
            guardar_detalle = detalleVentas(
                numero_orden=venta.numero_orden,  # Aquí asignamos el número de orden, no el objeto completo
                nombre_cliente=nombre_cliente,
                correo_cliente=correo_cliente,
                pedido=pedido,
                cantidad_productos=cantidad_productos,
            )

            guardar_detalle.save()

            return JsonResponse({"mensaje": "Detalle de venta guardado correctamente"})

        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al procesar el JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)


# class listar_finanzas(TemplateView):
#     template_name = "finanzas.html"

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         ventas = Ventas.objects.all()
#         labels = [venta.numero_orden for venta in ventas]
#         datos = [venta.total_precio_venta for venta in ventas]
#         context["labels"] = labels
#         context["datos"] = datos
#         return context


def finanzas_view(request):
    hoy = datetime.now()
    ayer = hoy - timedelta(days=1)
    # inicio_dia = timezone.make_aware(datetime.combine(hoy, datetime.min.time()), timezone.get_current_timezone())
    # fin_dia = inicio_dia + timedelta(days=1)

    print("dia de hoy: ", hoy)
    print("ayer: ", ayer)
    # Ventas por mes
    ventas_por_mes = (
        Ventas.objects.annotate(mes=TruncMonth("fecha_creacion"))
        .values("mes")
        .annotate(
            total_ventas=Sum("total_precio_venta"),
            total_productos=Sum("cantidad_productos"),
        )
        .order_by("mes")
    )

    # Comisiones por empleado con cantidad de ventas
    comisiones_por_empleado = (
        Ventas.objects.values("empleado")
        .annotate(
            total_ventas=Sum("total_precio_venta"),
            total_comisiones=Sum(
                F("total_precio_venta") * 0.10
            ),  # Asumiendo una comisión del 10%
            cantidad_ventas=Count("id"),
        )
        .order_by("empleado")
    )

    # Ventas de hoy
    ventas_hoy = Ventas.objects.filter(fecha_creacion=hoy)

    # Ventas de ayer
    ventas_ayer = Ventas.objects.filter(fecha_creacion=ayer)

    context = {
        "ventas_por_mes": ventas_por_mes,
        "comisiones_por_empleado": comisiones_por_empleado,
        "ventas_hoy": ventas_hoy,
        "ventas_ayer": ventas_ayer,
    }

    return render(request, "finanzas.html", context)



def generar_reporte_excel(request):
    # Crear un libro de trabajo (workbook)
    wb = Workbook()

    # Función para agregar datos a una hoja de trabajo con formato mejorado
    def agregar_datos(hoja, titulo, datos):
        hoja.title = titulo
        headers = datos[0].keys() if datos else []
        columnas = [get_column_letter(i + 1) for i in range(len(headers))]
        
        # Estilos para encabezados
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_border = Border(bottom=Side(border_style="thin"))

        # Escribir encabezados
        for col, header in zip(columnas, headers):
            cell = hoja[f"{col}1"]
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = header_border
            hoja.column_dimensions[col].width = max(len(str(header)) + 4, 12)

        # Estilos para datos
        data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Escribir datos
        for i, dato in enumerate(datos, start=2):
            for j, key in enumerate(headers):
                cell = hoja[f"{columnas[j]}{i}"]
                cell.value = dato[key]
                cell.alignment = data_alignment
                cell.border = data_border
                hoja.column_dimensions[columnas[j]].width = max(len(str(dato[key])) + 4, 12)

    # Ventas por mes con más información
    ventas_por_mes = Ventas.objects.annotate(
        mes=TruncMonth('fecha_creacion')
    ).values('mes').annotate(
        total_ventas=Sum('total_precio_venta'),
        total_productos=Sum('cantidad_productos'),
        promedio_venta=Sum('total_precio_venta') / Sum('cantidad_productos'),
        descuento_promedio=Sum('descuento') / Count('id'),
        empleados_distintos=Count('empleado', distinct=True)
    ).order_by('mes')

    # Ventas de hoy
    hoy = datetime.now()
    
    ventas_hoy = Ventas.objects.filter(fecha_creacion= hoy).values()

    # Todas las ventas
    todas_las_ventas = Ventas.objects.all().values()

    # Agregar datos a las hojas
    agregar_datos(wb.active, "Ventas por Mes", list(ventas_por_mes))
    agregar_datos(wb.create_sheet(title="Ventas Hoy"), "Ventas Hoy", list(ventas_hoy))
    agregar_datos(wb.create_sheet(title="Todas las Ventas"), "Todas las Ventas", list(todas_las_ventas))

    # Guardar el libro de trabajo en memoria y devolver la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openpyxl.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
    wb.save(response)

    return response
